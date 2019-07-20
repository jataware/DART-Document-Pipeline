from __future__ import print_function
import  openpyxl
import boto3, json
import os
from shutil import copyfile
from hashlib import sha256
from elasticsearch import Elasticsearch, RequestsHttpConnection
from requests_aws4auth import AWS4Auth
from bs4 import BeautifulSoup
from jsonschema import validate
import warnings
import re
import requests
import urllib3
import magic
import configparser
from PIL import Image 
import pytesseract 
import sys 
from pdf2image import convert_from_path
from datetime import datetime
import html2text
from tika import parser	
import PyPDF2
import threading
import time
from io import StringIO
import traceback
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.filterwarnings("ignore", category=PyPDF2.utils.PdfReadWarning)

TEMP_DOWNLOAD_PATH = '/tmp'
OUTPUT_LOCK  = threading.Lock()

# CONFIG

config = configparser.ConfigParser()
config.readfp(open(r'../config'))
DOC_TYPE = config.get('ES', 'doc_type')

# AWS CONFIG
AWS_PROFILE = config.get('AWS', 'profile')

# S3 CONFIG
S3_BASE_URL = config.get('S3', 'base_url')
S3_BASE_KEY = config.get('S3', 'base_key')
BUCKET_NAME = config.get('S3', 'bucket_name')

# ELASTIC SEARCH CONFIG
ES_INDEX = config.get('ES', 'index')
DOC_TYPE = config.get('ES', 'doc_type')
REGION = config.get('ES', 'region')
SERVICE = config.get('ES', 'service')
ES_HOST = config.get('ES', 'host')

# LOGGING
FAILURE_FILE = config.get('LOGGING', 'fail_file')

SPREADSHEET = 'Six_Twelve-Month_November_2019_Evaluation_Documents-Updated-6June2019.xlsx'
SHEET_NAMES = [
    'Six-Month Evaluation Documents',
    'Additional Six-Month Eval Docs',
    'Twelve-Month Eval Docs',
    'November 2019 SSudan Docs',
    'November 2019 Ethiopia Docs',
    'Luma-Provided Ethiopia Docs'
]

def extract_bs4(file_path, extracted_text):
    """
    Take in a file path of an HTML document and return its Beautiful Soup extraction
    https://www.crummy.com/software/BeautifulSoup/bs4/doc/
    """       
    try: 
        htmlFileObj = open(file_path, 'r')
        soup = BeautifulSoup(htmlFileObj, "lxml")
        # kill all script and style elements
        for script in soup(["script", "style"]):
            script.decompose()    # rip it out
        # get text
        text = soup.get_text()        
        # break into lines and remove leading and trailing space on each
        lines = (line.strip() for line in text.splitlines())
        # break multi-headlines into a line each
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        # drop blank lines
        bs4_extraction = '\n'.join(chunk for chunk in chunks if chunk)
        extracted_text['bs4'] = bs4_extraction
    except Exception:
        with OUTPUT_LOCK:
            print(f"Failed extracting BS4: {traceback.format_exc()}")

def convertPdfDatetime(pd):
    dtformat = "%Y%m%d%H%M%S"
    clean = pd.decode("utf-8").replace("D:","").split('-')[0].split('+')[0]
    return datetime.strptime(clean,dtformat)

def pdfminer_parse_pdfinfo(doc, fp):
    """
    Takes in pdfinfo from PDFMiner and a document and enriches the document
    with metadata fields
    """
    parser = PDFParser(fp)
    pdfdoc = PDFDocument(parser)
    info = pdfdoc.info[0]

    title = info.get('Title',None)
    date = info.get('CreationDate', info.get('created',None))
    author = info.get('Author',None)
    last_modified = info.get('ModDate',None)
    if title:
        doc['title'] = title.decode('utf-8')
    if date:
        doc['creation_date'] = {'date': convertPdfDatetime(date).strftime('%Y-%m-%dT%H:%M:%SZ')}
    if author:
        doc['source'] = {'author_name': author.decode('utf-8')}
    if last_modified:
        doc['modification_date'] = {'date': convertPdfDatetime(last_modified).strftime('%Y-%m-%dT%H:%M:%SZ')}
    return doc

def parse_pdfinfo(t_m, doc, file_path):
    """
    Takes in pdfinfo from tika and a document and enriches the document
    with metadata fields
    """
    title = t_m.get('title',None)
    date = t_m.get('Creation-Date',t_m.get('created',None))
    author = t_m.get('Author',None)	
    last_modified = t_m.get('Last-Modified',None)

    if title:
        doc['title'] = title
    if date:
        doc['creation_date'] = {'date': date}
    if author:
        doc['source'] = {'author_name': author}
    if last_modified:
        doc['modification_date'] = {'date': last_modified}
    return doc

def extract_tika(file_path, extracted_text, tm):	
    """	
    Take in a file path of a PDF and return its Tika extraction	
    https://github.com/chrismattmann/tika-python	
    	
    Returns: a tuple of (extracted text, extracted metadata)	
    """	
    try:
        tika_data = parser.from_file(file_path)	
        tika_extraction = tika_data.pop('content')	
        tika_metadata = tika_data.pop('metadata')
        extracted_text['tika'] = tika_extraction
        tm = tika_metadata
    except Exception:
        with OUTPUT_LOCK:
            print(f"Tika extraction failed: {traceback.format_exc()}") 

def extract_pypdf2(file_path, extracted_text):	
    """	
    Take in a file path of a PDF and return its PyPDF2 extraction	
    https://github.com/mstamy2/PyPDF2	
    """	
    try:
        pdfFileObj = open(file_path, 'rb')	
        pdfReader = PyPDF2.PdfFileReader(pdfFileObj)	
        page_count = pdfReader.numPages	
        pypdf2_extraction = ''	
        for page in range(page_count):	
            pageObj = pdfReader.getPage(page)	
            page_text = pageObj.extractText()	
            pypdf2_extraction += page_text
        extracted_text['pypdf2'] = pypdf2_extraction
    except Exception:
        with OUTPUT_LOCK:
            print(f"PyPDF2 extraction failed: {traceback.format_exc()}")

def extract_html2text(fp, extracted_text):
    try:
        extracted_text['htm2text'] = html2text.html2text(fp.read())
    except Exception:
        with OUTPUT_LOCK:
            print(f"Failed extracting html2text: {traceback.format_exc()}")

def extract_pytesseract(file_path, extracted_text):
    try:
        pages = convert_from_path(file_path, 500)
        out_text = ""
        image_counter = 1    
        for page in pages: 
            filename = f"{TEMP_DOWNLOAD_PATH}/page_{str(image_counter)}.jpg"
            page.save(filename, 'JPEG')     
            image_counter = image_counter + 1
        
        filelimit = image_counter-1
        
        for i in range(1, filelimit + 1): 
            filename = f"{TEMP_DOWNLOAD_PATH}/page_{str(i)}.jpg"            
            text = str(((pytesseract.image_to_string(Image.open(filename))))) 
            text = text.replace('-\n', '')         
            out_text += text

        extracted_text['pytesseract'] = out_text
    except Exception:
        with OUTPUT_LOCK:
            print(f"PyTesseract extraction failed: {traceback.format_exc()}")

def extract_pdfminer(fp, extracted_text, pages=None):
    try:
        if not pages:
            pagenums = set()
        else:
            pagenums = set(pages)

        output = StringIO()
        manager = PDFResourceManager()
        converter = TextConverter(manager, output, laparams=LAParams())
        interpreter = PDFPageInterpreter(manager, converter)

        for page in PDFPage.get_pages(fp, pagenums):
            interpreter.process_page(page)
        converter.close()
        text = output.getvalue()
        output.close
        extracted_text['pdfminer'] = text
    except Exception as e:
        with OUTPUT_LOCK:
            print(f"PDFMiner extraction failed: {traceback.format_exc()}")


def parse_document(file_path, category, source_url):
    """
    Take in the full path to a file and perform appropriate text extrraction
    as well as metadata enrichment (if a PDF, using pdfinfo fields)
    """
    threads = []
    file_name = file_path.split('/')[-1]
    mime = magic.Magic(mime=True)
    file_ext = source_url.split('.')[-1].split('?')[0].lower()
    if 'pdf' in file_ext:
        file_type = 'pdf'
    else:
        file_type = mime.from_file(file_path).split('/')[-1]
    
    # sha256 hash the raw contents of the file to generate a UUID
    with open(file_path, 'rb') as fp:
        _id = sha256(fp.read()).hexdigest()
        
        doc = {'_id': _id,
            'file_name': file_name, 
            'file_type': file_type,
            'category': category,
            'source_url': source_url}
        
        extracted_text = {}
        
        if 'pdf' in file_type or 'xml' in file_type:
            doc['file_type'] = 'pdf'
            tika_metadata = None
            extract_tika_thread = threading.Thread(target=extract_tika, args=(file_path, extracted_text, tika_metadata))
            threads.append(extract_tika_thread)

            extract_pypdf2_thread = threading.Thread(target=extract_pypdf2, args=(file_path, extracted_text))
            threads.append(extract_pypdf2_thread)

            extract_pdfminer_thread = threading.Thread(target=extract_pdfminer, args=(file_path, extracted_text))
            threads.append(extract_pdfminer_thread)
            
            extract_pytesseract_thread = threading.Thread(target=extract_pytesseract, args=(file_path, extracted_text))
            threads.append(extract_pytesseract_thread)
        elif 'html' in file_type or 'text' in file_type:
            doc['file_type'] = 'html'
            extract_bs4_thread = threading.Thread(target=extract_bs4, args=(file_path, extracted_text))
            threads.append(extract_bs4_thread)
            extract_html2text_thread = threading.Thread(target=extract_html2text, args=(fp, extracted_text))
            threads.append(extract_html2text_thread)
        else:
            raise ValueError("*** Could not find extractor for "+file_name+" with mime type - "+file_type)
        
        for thread in threads:
            thread.start()

        for thread in threads:
            thread.join()
        try:
            doc = parse_pdfinfo(tika_metadata, doc, fp)
        except Exception:
            print(f"Error retrieving PDFINFO --- {traceback.format_exc()}")
        doc['extracted_text'] = extracted_text

        # This add_periods method is used with both html2text and pdfminer text extraction used by UAZ
        doc = add_periods(doc)
        validate_extracted_text(doc['extracted_text'], file_name, file_type)
        return doc

def validate_extracted_text(extracted_text, file_name, file_type):
    bs4_len = len(extracted_text.get('bs4') or '')
    pdfminer_len = len(extracted_text.get('pdfminer') or '')
    tika_len = len(extracted_text.get('tika') or '')
    pypdf2_len = len(extracted_text.get('pypdf2') or '')
    html2text_len = len(extracted_text.get('html2text') or '')
    pytesseract_len = len(extracted_text.get('pytesseract') or '')
    if bs4_len < 500 and pdfminer_len < 500 and pytesseract_len < 500 and html2text_len < 500 and tika_len < 500 and pypdf2_len < 500:
        ERRORS.append("*** Error extracted_text for "+file_name+" with type "+file_type+" is less than 500 chars - "+json.dumps(extracted_text))
        raise ValueError('ERRORS --- ' + json.dumps(ERRORS))

def slugify(value):
    return ''.join([c for c in value if c.isalpha() or c.isdigit() or c ==' ' or c == '.']).rstrip()

def get_filename(req, url):
    header = req.headers
    content_type = header.get('content-type')
    ext = content_type.split(';')[0].split('/')[-1]
    return f"{url.split('/')[-1][:225].split('.')[0]}.{ext}"

def regex_periods(text):
    res = re.sub(r'\s*\n\s*\n\s*', '.\n\n', text)
    return re.sub(r'\.\.', '.', res)

def add_periods(doc):
    pdfminer_text = doc['extracted_text'].get('pdfminer', None)
    html2text_text = doc['extracted_text'].get('html2text', None)
    if pdfminer_text:
        doc['extracted_text']['pdfminer'] = regex_periods(pdfminer_text)
    if html2text_text:
        doc['extracted_text']['html2text'] = regex_periods(html2text_text)
    return doc


def connect_to_es():
    session = boto3.Session(region_name=REGION, profile_name='wmuser')
    credentials = session.get_credentials()
    credentials = credentials.get_frozen_credentials()
    access_key = credentials.access_key
    secret_key = credentials.secret_key
    token = credentials.token

    aws_auth = AWS4Auth(
        access_key,
        secret_key,
        REGION,
        SERVICE,
        session_token=token
    )
    
    return Elasticsearch(
        hosts = [{'host': ES_HOST, 'port': 443}],
        http_auth=aws_auth,
        use_ssl=True,
        verify_certs=True,
        connection_class=RequestsHttpConnection
    )

def connect_to_s3():
    session = boto3.Session(profile_name=AWS_PROFILE)
    s3 = session.resource("s3")
    return boto3.client("s3")

def main():
    es = connect_to_es()
    s3_client = connect_to_s3()
    if not es.indices.exists(ES_INDEX):
        es.indices.create(ES_INDEX)
        print(f"Created ES index: {ES_INDEX}")
    
    # Ensure we are connected
    print(json.dumps(es.info(), indent=2))
    
    schema = json.loads(open("../document-schema.json").read())
    book = openpyxl.load_workbook(SPREADSHEET)
    failed = openpyxl.Workbook()
    for name in SHEET_NAMES:
        sheet = book[name]
        failed.create_sheet(name)
        for row in range(2, sheet.max_row):
            try:
                doc_name = sheet[f"A{row}"].value
                url_path = sheet[f"D{row}"].value.strip()
                doc_date = sheet[f"B{row}"].value
                query = {
                    "query": {
                        "match" : {
                            "source_url.keyword" : url_path
                        }
                    }
                }            
                
                es_count = es.count(index=ES_INDEX, body=query)['count']
                if es_count < 1:
                    print(f"Processing - {doc_name}")
                    print("Downloading - %s" % (url_path,))
                    r = requests.get(url_path, verify=False, stream=True, allow_redirects=True)
                    r.raw.decode_content = True
                    filename = f"{TEMP_DOWNLOAD_PATH}/{slugify(get_filename(r, url_path))}"
                    open(filename, 'wb').write(r.content)
                    s3_key = f"{S3_BASE_KEY}{TEMP_DOWNLOAD_PATH}/DEV/{filename.split('/')[-1]}"
                    s3_uri = f"{S3_BASE_URL}/{s3_key}"
                    s3_client.upload_file(filename, BUCKET_NAME, s3_key)
                    
                    title = doc_name
                    category = name
                    source_url = url_path
                    creation_date = doc_date
                    doc = parse_document(filename, category, source_url)
                    doc['stored_url'] = s3_uri
                    
                    validate(instance=doc, schema=schema)
                        
                    es.index(index=ES_INDEX, doc_type=DOC_TYPE, id=doc.pop('_id'), body=doc)
                    print(f"Finished processing row# {row} out of {sheet.max_row} in sheet {name}")
                    failed[name].delete_rows(row, 1)
            except Exception:
                print(f"Failed processing row# {row} out of {sheet.max_row} in sheet {name} -- {traceback.format_exc()}")
                failed[name][f"A{row}"] = sheet[f"A{row}"].value
                failed[name][f"B{row}"] = sheet[f"B{row}"].value
                failed[name][f"C{row}"] = sheet[f"C{row}"].value
                failed[name][f"D{row}"] = sheet[f"D{row}"].value
                failed[name][f"E{row}"] = traceback.format_exc().replace('"', '\\"')
                failed.save(FAILURE_FILE)            

if __name__ == '__main__':
    main()
